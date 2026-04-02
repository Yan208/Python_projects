# main.py

from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.clock import Clock
from kivy.core.window import Window
from kivy.metrics import dp
from openpyxl import load_workbook
import random
import os

# Загрузчик слов из Excel
class WordsLoader:
    def __init__(self, filename):
        self.filename = filename

    def get_words(self):
        book = load_workbook(self.filename)
        sheet = book['Лист1']
        words = []
        for row in range(1, sheet.max_row + 1):
            value = sheet[f'A{row}'].value
            if value:
                words.append(value)
        return words

# Экран запоминания (Saving)
class SavingScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.timer = 10
        self.timer_event = None
        self.words_container = None
        self.timer_label = None

        # Основной вертикальный layout
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        self.add_widget(layout)

        # Заголовок
        layout.add_widget(Label(text="Запомните список слов!",
                                size_hint_y=0.1,
                                font_size=dp(18)))

        # ScrollView для списка слов (на случай длинного списка)
        scroll = ScrollView(size_hint_y=0.6)
        self.words_container = GridLayout(cols=1, spacing=5, size_hint_y=None)
        self.words_container.bind(minimum_height=self.words_container.setter('height'))
        scroll.add_widget(self.words_container)
        layout.add_widget(scroll)

        # Метка таймера
        self.timer_label = Label(text=f"Осталось {self.timer}",
                                 size_hint_y=0.1,
                                 font_size=dp(16))
        layout.add_widget(self.timer_label)

        # Кнопка "Запомнил"
        btn = Button(text="Запомнил",
                     size_hint_y=0.1,
                     font_size=dp(16))
        btn.bind(on_press=self.hide_words)
        layout.add_widget(btn)

    def on_enter(self):
        app = App.get_running_app()
        self.words_container.clear_widgets()
        for i, word in enumerate(app.words, 1):
            # Метка с фиксированной высотой и увеличенным шрифтом
            lbl = Label(text=f"{i}. {word}",
                        halign='left',
                        size_hint_y=None,
                        height=dp(40),
                        font_size=dp(16))
            self.words_container.add_widget(lbl)

        self.timer = 10
        self.timer_label.text = f"Осталось {self.timer}"
        if self.timer_event:
            self.timer_event.cancel()
        self.timer_event = Clock.schedule_interval(self.tick, 1)

    def tick(self, dt):
        self.timer -= 1
        self.timer_label.text = f"Осталось {self.timer}"
        if self.timer <= 0:
            self.timer_event.cancel()
            self.hide_words()

    def hide_words(self, *args):
        if self.timer_event:
            self.timer_event.cancel()
        self.manager.current = 'exec'

# Экран ввода и проверки (Exec)
class ExecScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.entries = []
        self.result_labels = []

        # Основной layout – ScrollView, чтобы можно было прокрутить всё, включая кнопку
        scroll = ScrollView()
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10, size_hint_y=None)
        layout.bind(minimum_height=layout.setter('height'))
        scroll.add_widget(layout)
        self.add_widget(scroll)

        # Заголовок
        layout.add_widget(Label(text="Вспомни и запиши слова!",
                                size_hint_y=None,
                                height=dp(40),
                                font_size=dp(18)))

        # Контейнер для полей ввода (будет заполнен в on_enter)
        self.input_container = GridLayout(cols=1, spacing=10, size_hint_y=None)
        self.input_container.bind(minimum_height=self.input_container.setter('height'))
        layout.add_widget(self.input_container)

        # Кнопка проверки (добавляется в конец layout, всегда будет внизу прокрутки)
        btn = Button(text="Проверить",
                     size_hint_y=None,
                     height=dp(50),
                     font_size=dp(16))
        btn.bind(on_press=self.check_words)
        layout.add_widget(btn)

        # Настройка поведения при появлении клавиатуры
        Window.softinput_mode = 'below_target'   # поднимаем окно, чтобы поле было видно

    def on_enter(self):
        app = App.get_running_app()
        self.input_container.clear_widgets()
        self.entries.clear()
        self.result_labels.clear()

        for i, word in enumerate(app.words, 1):
            # Горизонтальный ряд: номер + поле + результат
            row = BoxLayout(orientation='horizontal',
                            size_hint_y=None,
                            height=dp(48))  # увеличенная высота строки
            # Номер
            row.add_widget(Label(text=f"{i}.", size_hint_x=0.1, font_size=dp(16)))
            # Поле ввода
            entry = TextInput(size_hint_x=0.6,
                              multiline=False,
                              font_size=dp(16))
            row.add_widget(entry)
            # Метка результата
            result_lbl = Label(text="", size_hint_x=0.3, font_size=dp(16))
            row.add_widget(result_lbl)

            self.input_container.add_widget(row)
            self.entries.append(entry)
            self.result_labels.append(result_lbl)

    def check_words(self, instance):
        app = App.get_running_app()
        for i, entry in enumerate(self.entries):
            user_word = entry.text.strip()
            original = app.words[i]
            if user_word == original:
                self.result_labels[i].text = "Ок!"
            else:
                self.result_labels[i].text = "Не Ок!"

# Главное приложение
class WordsApp(App):
    def build(self):
        # Загружаем слова (относительный путь относительно расположения скрипта)
        script_dir = os.path.dirname(os.path.abspath(__file__))
        words_path = os.path.join(script_dir, 'words.xlsx')
        loader = WordsLoader(words_path)
        self.words = loader.get_words()
        random.shuffle(self.words)

        # Менеджер экранов
        sm = ScreenManager()
        sm.add_widget(SavingScreen(name='saving'))
        sm.add_widget(ExecScreen(name='exec'))
        sm.current = 'saving'
        return sm

if __name__ == '__main__':
    WordsApp().run()
    